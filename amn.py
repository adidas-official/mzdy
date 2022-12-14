# amn.py
# -------------+
# Automatizace |
# Mzdovych     |
# Nakladu      |
# -------------+
# v4
# author: Zdenek Frydryn
# created for: Bereko s.r.o., Drogerie Fiala s.r.o.,
# Description: Automatizace vyplnovani mezd do xlsx tabulek pro urad prace a internich tabulek.

# TOOD:
# - add ozp status from `PRACOV.CSV` - DONE
# - open src folder and output folder for manual copying of checked file
# --------------------------------------------------------------------------------

import logging
import openpyxl
import msoffcrypto
from pathlib import Path
from openpyxl.utils import column_index_from_string, get_column_letter

from months_cz import months_cz
from io import BytesIO
import json
import tkinter as tk
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText
from datetime import datetime
from functions import load_ins_codes, add_new, update_ins, delete_record, item_selected, activate_tab, set_dir, \
    show_banner, set_datas, rename_tab, open_output, delete_tab, main_window, prepare_input, check_new_ppl
from os import remove, system
from shutil import copy

HOME = Path.home() / 'amn'
STRUCTURE = 'structure.json'

if not HOME.exists():
    HOME.mkdir()

if not Path(HOME / 'structure.json').exists():
    copy(Path.cwd() / 'structure.json', HOME / 'structure.json')

STRUCTURE = HOME / 'structure.json'
NEW_PPL = HOME / 'new_ppl.txt'


logging.basicConfig(level=logging.INFO, filename='log.log', filemode='w',
                    format='%(levelname)s - %(asctime)s - %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S')

current_month = months_cz[(datetime.now().month - 2) % 12]

with open(STRUCTURE, 'r', encoding='cp1250') as structure_data:
    last_data = json.load(structure_data)


def new_company(place):
    new_name = new_company_name_entry.get()
    companies.append(new_name)
    if new_name not in last_data:
        last_data[new_name] = {'input_data': '', 'newguys': '', 'src_file_up': '', 'src_file_loc': '', 'output': ''}
        make_frame(new_name, place)

        with open(STRUCTURE, 'w') as outfile:
            outfile.write(json.dumps(last_data, indent=4))

    else:
        print(f'{new_name} already exists')
        txt.insert('1.0', f'Nazev firmy `{new_name}` uz je pouzit. Vyberte jiny nazev.')


companies = []


def amn(month_name, text_field):
    with open(STRUCTURE, 'r', encoding='cp1250') as jdata:
        data = json.load(jdata)

    text_field.delete('1.0', tk.END)
    month = months_cz.index(month_name) + 1

    if Path(NEW_PPL).exists():
        remove(NEW_PPL)

    for c_name, input_output in data.items():

        if c_name not in companies:
            continue

        new_ppl_file = open(NEW_PPL, 'a', encoding='cp1250')
        new_ppl_file.write(c_name + '\n')

        text_field.insert(tk.END, f'|- Vyplnuji {c_name}\n')
        text_field.insert(tk.END, '=' * 44 + '\n')
        logging.info(f'Vyplnuji {c_name}')

        wb = openpyxl.load_workbook(input_output['src_file_up'])
        ws = wb["1) Úvodní list"]
        date_formated = datetime.now().strftime("%d.%m.%Y")
        ws.cell(row=30, column=column_index_from_string('B')).value = 'Dalovicích'
        ws.cell(row=30, column=column_index_from_string('E')).value = date_formated

        # fill in company details on first page, if it is not filled in already
        # companyfullnamecell = ws.cell(row=9, column=column_index_from_string('E'))
        # companyidcell = ws.cell(row=10, column=column_index_from_string('E'))
        # companyownercell = ws.cell(row=30, column=column_index_from_string('H'))
        #
        # if not companyfullnamecell.value:
        #     companyfullnamecell.value = input_output['fullcname']
        # if not companyidcell.value:
        #     companyidcell.value = input_output['ico']
        # if not companyownercell.value:
        #     companyownercell.value = input_output['cowner']


        if month <= 3:
            ws.cell(row=6, column=4).value = 1
        elif 3 < month <= 6:
            ws.cell(row=6, column=4).value = 2
        elif 6 < month <= 9:
            ws.cell(row=6, column=4).value = 3
        else:
            ws.cell(row=6, column=4).value = 4

        employees_up, employees_inter = prepare_input(input_output['input_data'], c_name)
        new_or_dead_p, new_or_dead_loc = check_new_ppl(input_output['newguys'])

        #for n in new_or_dead_p.items():
        #    print(n)           # ('7308021908', {'VstupDoZam': '21.06.2022', 'UkonceniZam': '', 'DuchOd': '', 'TypDuch': '', 'Jmeno': "'Paulus Bohuslav'", 'Kat': "'DPP'", 'Kod': "'Admin'"})

        # open mzdy UP table, go through each name in data and check if it is in table
        sheets = ['2) jmenný seznam', '3) nákl. prov. z. a prac. a.']
        col_letter_id = column_index_from_string('D')

        if month % 3 == 1:
            col_letter_pay = {sheets[0]: 11, sheets[1]: 9}
        elif month % 3 == 2:
            col_letter_pay = {sheets[0]: 16, sheets[1]: 10}
        else:
            col_letter_pay = {sheets[0]: 21, sheets[1]: 11}

        ids_in_xlsx = {}
        last_rows = []

        for sheet in sheets:
            progress['value'] = 0
            root.update_idletasks()
            ws = wb[sheet]
            last_row = 13
            # ids_in_xlsx.setdefault(sheet, {})
            ids_in_xlsx[sheet] = {}

            for i in range(100):
                row = i + 13
                id_of_person = ws.cell(row=row, column=col_letter_id).value
                # make list of ids in xlsx file
                if id_of_person:
                    ids_in_xlsx[sheet].setdefault(str(id_of_person).replace('/', ''), row)
                    last_row = row

            last_rows.append(last_row)

        for emp_id, emp_data in employees_up.items():
            # print(emp_id, type(emp_id))
            progress['value'] += 100 / (len(employees_up))
            found = False
            for sheet_data in ids_in_xlsx.values():
                if str(emp_id) in sheet_data:
                    found = True
            if not found:  # in mzdy.csv NOT in xlsx, emp_id is from csv file
                # print(emp_id, type(emp_id), emp_data)  # 0402186950 {'first name': 'Michal', 'last name': 'Kratochvíl', 'ins code': '111', 'cat': 'U', 'payment expenses': 4800}

                logging.info(f"Novy zamestnanec {emp_data['first name']} {emp_data['last name']}")
                text_field.insert(tk.END, f"|- Novy zam. {emp_data['first name']} {emp_data['last name']} ")
                text_field.see(tk.END)
                root.update_idletasks()
                if emp_data['cat'] == 'INV':
                    text_field.insert(tk.END, f"-> {sheets[0][:3]}\n")
                    logging.info(f"Pridavam do listu {sheets[0]}")
                    ws = wb[sheets[0]]
                    ws.cell(row=last_rows[0] + 1, column=2).value = emp_data['last name']
                    ws.cell(row=last_rows[0] + 1, column=3).value = emp_data['first name']
                    ws.cell(row=last_rows[0] + 1, column=4).value = emp_id
                    ws.cell(row=last_rows[0] + 1, column=7).value = emp_data['ins code']
                    ws.cell(row=last_rows[0] + 1, column=col_letter_pay[ws.title]).value = emp_data['payment expenses']

                    if emp_id in new_or_dead_p:
                        ws.cell(row=last_rows[0] + 1, column=5).value = new_or_dead_p[emp_id]['VstupDoZam']
                        ws.cell(row=last_rows[0] + 1, column=6).value = new_or_dead_p[emp_id]['UkonceniZam']
                        ws.cell(row=last_rows[0] + 1, column=8).value = new_or_dead_p[emp_id]['DuchOd']
                        ws.cell(row=last_rows[0] + 1, column=col_letter_pay[ws.title] - 1).value = new_or_dead_p[emp_id]['TypDuch']
                        ws.cell(row=last_rows[0] + 1, column=col_letter_pay[ws.title]).value = emp_data['payment expenses']

                    last_rows[0] += 1
                elif emp_data['cat'] == 'U':
                    text_field.insert(tk.END, ":Ucen\n")
                    logging.info("Ucen")
                    continue

                else:
                    text_field.insert(tk.END, f"-> {sheets[1][:3]}\n")
                    logging.info(f"Pridavam do listu {sheets[1]}")
                    ws = wb[sheets[1]]
                    ws.cell(row=last_rows[1] + 1, column=2).value = emp_data['last name']
                    ws.cell(row=last_rows[1] + 1, column=3).value = emp_data['first name']
                    ws.cell(row=last_rows[1] + 1, column=4).value = str(emp_id)[:6] + '/' + str(emp_id)[6:]
                    ws.cell(row=last_rows[1] + 1, column=col_letter_pay[ws.title]).value = emp_data['payment expenses']
                    ws.cell(row=last_rows[1] + 1, column=5).value = '-\'\'-'
                    ws.cell(row=last_rows[1] + 1, column=6).value = 'PA'
                    ws.cell(row=last_rows[1] + 1, column=7).value = '100%'

                    last_rows[1] += 1

                new_ppl_file.write(emp_data['last name'] + ' ' + emp_data['first name'] + ' ' + str(emp_data['payment expenses']) + '\n')

        text_field.insert(tk.END, "=" * 44 + '\n')
        logging.info("-" * 44)

        for sheet_name, data in ids_in_xlsx.items():
            progress['value'] = 0
            ws = wb[sheet_name]

            for id_num, row_num in data.items():
                # print(id_num, type(id_num))
                progress['value'] += 100 / (len(data.items()))
                if id_num in employees_up:
                    text_field.insert(
                        tk.END,
                        f"|- {employees_up[id_num]['first name']} {employees_up[id_num]['last name']}:{employees_up[id_num]['payment expenses']}\n")
                    logging.info(
                        f"Vyplnuji vyplatu pro {employees_up[id_num]['first name']} {employees_up[id_num]['last name']} v listu {sheet_name}:{employees_up[id_num]['payment expenses']}")
                    # zadat plat za tento mesic
                    ws.cell(row=row_num, column=col_letter_pay[sheet_name]).value = employees_up[id_num][
                        'payment expenses']
                    if sheet_name == sheets[0]:
                        if col_letter_pay[sheet_name] == 21:  # last month
                            ozp_status = ws.cell(row=row_num, column=col_letter_pay[sheet_name] - 6).value
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] - 1).value = ozp_status

                            # add formula if salary is > 0
                            if employees_up[id_num]['payment expenses']:
                                ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 12).value = f'=14200-{get_column_letter(col_letter_pay[sheet_name]+1)}{row_num}'
                            else:
                                ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 12).value = ''
                        elif col_letter_pay[sheet_name] == 11:  # first month
                            # delete salary in next 2 months
                            ozp_status = ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 9).value
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] - 1).value = ozp_status
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 4).value = ''
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 5).value = ''
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 9).value = ''
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 10).value = ''

                            # add formula if salary is > 0
                            if employees_up[id_num]['payment expenses']:
                                ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 14).value = f'=14200-{get_column_letter(col_letter_pay[sheet_name]+1)}{row_num}'
                            else:
                                ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 14).value = ''

                        elif col_letter_pay[sheet_name] == 16:  # second month
                            ozp_status = ws.cell(row=row_num, column=col_letter_pay[sheet_name] - 6).value
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] - 1).value = ozp_status
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 4).value = ''
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 5).value = ''

                            # add formula if salary is > 0
                            if employees_up[id_num]['payment expenses']:
                                ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 13).value = f'=14200-{get_column_letter(col_letter_pay[sheet_name]+1)}{row_num}'
                            else:
                                ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 13).value = ''
                    elif sheet_name == sheets[1]:
                        if col_letter_pay[sheet_name] == 9:  # first month
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 1).value = ''
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 2).value = ''
                        elif col_letter_pay[sheet_name] == 10:  # second month
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 1).value = ''

                else:
                    text_field.insert(tk.END,
                                      f"|- Zamestnanec {id_num} v listu {sheet_name} nema vyplnenou vyplatu. Mazu pole status\n")
                    logging.info(f"Zamestnanec {id_num} v listu {sheet_name} nema vyplnenou vyplatu. Mazu pole status")
                    # smazat v xlsx status pro tento mesic
                    if sheet_name == sheets[0]:
                        ws.cell(row=row_num, column=col_letter_pay[sheet_name] - 1).value = ''
                        ws.cell(row=row_num, column=col_letter_pay[sheet_name]).value = ''
                        if col_letter_pay[sheet_name] == 21:
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 12).value = ''
                        elif col_letter_pay[sheet_name] == 11:
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 14).value = ''
                        else:
                            ws.cell(row=row_num, column=col_letter_pay[sheet_name] + 13).value = ''

                root.update_idletasks()

        save_loc = Path(input_output['output']) / f'temp-{c_name}-up.xlsx'
        wb.save(save_loc)
        text_field.insert(tk.END, f"{c_name} vyplneno.\n")
        text_field.insert(tk.END, "=" * 44 + "\n")
        logging.info(f"{c_name} vyplneno.\n")

        if 'src_file_loc' in input_output.keys():
            if Path(input_output['src_file_loc']).exists():
                try:
                    decrypted_wb = BytesIO()
                    with open(input_output['src_file_loc'], 'rb') as f:
                        office_file = msoffcrypto.OfficeFile(f)
                        office_file.load_key(password='13881744')
                        office_file.decrypt(decrypted_wb)

                    wb = openpyxl.load_workbook(filename=decrypted_wb)
                except (UnboundLocalError, msoffcrypto.exceptions.FileFormatError):
                    wb = openpyxl.load_workbook(input_output['src_file_loc'])

                text_field.insert(tk.END, 'Vyplnuji data do interni tabulky\n')
                logging.info('Vyplnuji data do interni tabulky')
                fare_offset = 4


                for sheet in wb.sheetnames[:8]:
                    last_row = 0
                    progress['value'] = 0
                    root.update_idletasks()
                    if sheet == 'Úřad práce':
                        fare_offset = 2
                    text_field.insert(tk.END, '\n+ ' + sheet + '\n')
                    logging.info(sheet)
                    ws = wb[sheet]
                    month_col = ''
                    for col in range(2, column_index_from_string('CI')):
                        cell = ws.cell(row=1, column=col)

                        if not type(cell).__name__ == 'MergedCell' and cell.value:
                            if cell.value == month_name:
                                month_col = cell.column
                                break
                        else:
                            continue

                    for i in range(3, 203):
                        progress['value'] += 0.5
                        cell = ws.cell(row=i, column=2)
                        cell_val = cell.value
                        if cell_val:
                            #for n in new_or_dead_p.items():
                            #    print(n)   # ('7308021908', {'VstupDoZam': '21.06.2022', 'UkonceniZam': '', 'DuchOd': '', 'TypDuch': '', 'Jmeno': "'Paulus Bohuslav'", 'Kat': "'DPP'", 'Kod': "'Admin'"})


                            if cell_val == '[ENDBLOCK]':
                                break

                            cell_val = cell_val[:20]

                            if cell_val in employees_inter:

                                text_field.insert(tk.END, f"|- Vyplnuji {cell_val}: {employees_inter[cell_val][0]}\n")
                                logging.info(f"|- Vyplnuji {cell_val}: {employees_inter[cell_val][0]}")
                                ws.cell(row=i, column=month_col).value = employees_inter[cell_val][0]
                                ws.cell(row=i, column=month_col + fare_offset).value = employees_inter[cell_val][1]



                        root.update_idletasks()
                    # ADDING NEW PEOPLE
                    # pridat na konec, nechat v xlsx vzdy misto pro dalsi lidi. Takze zpet pridat plno prazdnych radku. super.

                    logging.info("")

                save_loc = Path(input_output['output']) / f'temp-{c_name}-loc.xlsx'
                wb.save(save_loc)
        new_ppl_file.write('\n')
        new_ppl_file.close()
        system(str(NEW_PPL))

    text_field.insert(tk.END, 'DONE')
    text_field.see('end')
    logging.info("DONE")


def make_frame(company_name, place):
    if company_name in last_data:
        file_paths = last_data[company_name]
    else:
        file_paths = False

    company_frame = ttk.Frame(tabs)
    company_frame.grid_columnconfigure(0, weight=1)
    company_frame.grid_columnconfigure(1, weight=1)
    company_frame.grid_columnconfigure(2, weight=1)
    company_frame.grid_columnconfigure(3, weight=1)

    active = tk.StringVar()
    check_box = ttk.Checkbutton(
        company_frame,
        text=company_name,
        onvalue=1,
        offvalue=0,
        variable=active,
        command=lambda state=active: activate_tab(
            root,
            tabs.select(),
            state,
            start_btn,
            company_name,
            companies
        )
    )
    check_box.grid(row=0, column=0)

    src_label = ttk.Label(company_frame, text='ZDROJ')
    src_label.grid(row=1, column=0)

    up_label_in = ttk.Label(company_frame, text='pro u.p.')
    up_label_in.grid(row=2, column=0)

    up_btn_in = ttk.Button(
        company_frame,
        text='Vyberte soubor',
        width=20,
        state='disabled',
        command=lambda: set_datas(
            root,
            company_name,
            last_data,
            up_btn_in,
            tabs.select(),
            [('Spreadsheets', '*.xlsx')],
            'src_file_up'
        )
    )

    if file_paths:
        up_btn_in['text'] = Path(file_paths['src_file_up']).name

    up_btn_in.grid(row=3, column=0)

    inter_label_in = ttk.Label(company_frame, text='interni')
    inter_label_in.grid(row=4, column=0)

    inter_btn_in = ttk.Button(
        company_frame,
        width=20,
        state='disabled',
        command=lambda: set_datas(
            root,
            company_name,
            last_data,
            inter_btn_in,
            tabs.select(),
            [('Spreadsheets', '*.xlsx')],
            'src_file_loc'
        )
    )

    if file_paths:
        inter_btn_in['text'] = Path(file_paths['src_file_loc']).name

    inter_btn_in.grid(row=5, column=0)

    rename_tab_btn = ttk.Button(
        company_frame,
        text='Prejmenovat',
        width=20,
        command=lambda: rename_tab(
            root,
            companies,
            last_data,
            tabs.select(),
            tabs
        )
    )

    rename_tab_btn.grid(row=8, column=0)

    delete_tab_btn = ttk.Button(
        company_frame,
        text='Smazat',
        width=20,
        command=lambda: delete_tab(
            last_data,
            companies,
            tabs
        )
    )

    delete_tab_btn.grid(row=8, column=1)

    output_label = ttk.Label(company_frame, text='VYSTUP')
    output_label.grid(row=1, column=1)

    output_open = ttk.Button(
        company_frame,
        text='Otevrit vystup',
        command=lambda: open_output(company_name, last_data),
        state='disabled'
    )
    output_open.grid(row=5, column=1)

    up_btn_out = ttk.Button(
        company_frame,
        text='Vyberte',
        width=20,
        state='disabled',
        command=lambda: set_dir(
            root,
            company_name,
            up_btn_out,
            tabs.select(),
            last_data
        )
    )

    if file_paths:
        up_btn_out['text'] = Path(file_paths['output']).name

    up_btn_out.grid(row=3, column=1)

    data_label = ttk.Label(company_frame, text='Data')
    data_label.grid(row=0, column=2)

    data_btn = ttk.Button(
        company_frame,
        text='Vyberte data',
        width=20,
        state='disabled',
        command=lambda: set_datas(
            root,
            company_name,
            last_data,
            data_btn,
            tabs.select(),
            [('CSV', '*.csv')],
            'input_data'
        )
    )

    if file_paths:
        data_btn['text'] = Path(file_paths['input_data']).name

    data_btn.grid(row=0, column=3)

    newguys_btn = ttk.Button(
        company_frame,
        text='Vyberte data',
        width=20,
        state='disabled',
        command=lambda: set_datas(
            root,
            company_name,
            last_data,
            newguys_btn,
            tabs.select(),
            [('CSV', '*.csv')],
            'newguys'
        )
    )

    if file_paths:
        newguys_btn['text'] = Path(file_paths['newguys']).name

    newguys_btn.grid(row=0, column=4)

    ins_groups_label = ttk.Label(company_frame, text='POJISTOVNY')
    ins_groups_label.grid(row=1, column=2)

    tree_label = ttk.Label(company_frame, text='kody zdravotnich pojistoven v ucto2000')
    tree_label.grid(row=2, column=2, columnspan=3)
    tree_cols = ('ucto_code', 'ins_code', 'ins_name')
    tree = ttk.Treeview(company_frame, columns=tree_cols, show='headings')
    tree.column('ucto_code', width=70)
    tree.column('ins_code', width=100)
    tree.heading('ucto_code', text='Ucto2000')
    tree.heading('ins_code', text='kod pojistovny')
    tree.heading('ins_name', text='nazev pojistovny')

    load_ins_codes(company_name, tree)

    tree.bind('<<TreeviewSelect>>',
              lambda event='<<TreeviewSelect>>', my_tree=tree: item_selected(root, event, my_tree))
    tree.grid(row=3, column=2, columnspan=3, rowspan=3)
    scrollbar = ttk.Scrollbar(company_frame, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscroll=scrollbar.set, height=5)

    code_ucto_btn = ttk.Entry(company_frame, state='disabled')
    code_ucto_btn.grid(row=7, column=2)

    ins_code_btn = ttk.Entry(company_frame, state='disabled')
    ins_code_btn.grid(row=7, column=3)

    ins_name_btn = ttk.Entry(company_frame, state='disabled')
    ins_name_btn.grid(row=7, column=4)

    add_button = ttk.Button(
        company_frame,
        text="Pridat",
        state='disabled',
        command=lambda: add_new(root, company_name, txt, tabs.select()))

    add_button.grid(row=8, column=2)

    update = ttk.Button(
        company_frame,
        text="Prepsat",
        state='disabled',
        command=lambda my_tree=tree: update_ins(root, company_name, tabs.select(), my_tree))

    update.grid(row=8, column=3)

    delete_btn = ttk.Button(
        company_frame,
        text="Smazat",
        state='disabled',
        command=lambda my_tree=tree: delete_record(company_name, my_tree))

    delete_btn.grid(row=8, column=4)

    for w in company_frame.winfo_children():
        w.grid(padx=5, pady=5, sticky='NWSE')
    scrollbar.grid(row=3, column=5, sticky='NS', rowspan=3)

    tabs.insert(place, company_frame, text=company_name)


root = tk.Tk()
style = ttk.Style(root)
style.theme_use('clam')

s = ttk.Style()
s.configure("Frames.TFrame", background='#EEE')

root.grid_columnconfigure(0, weight=1)

main_window(root, 725, 540)
opts = {'padx': 20, 'sticky': 'NSWE'}

top_frame = ttk.Frame(root, style="Frames.TFrame")

top_frame.grid_columnconfigure(0, weight=4)
top_frame.grid_columnconfigure(1, weight=1)
top_frame.grid_columnconfigure(2, weight=24)

chosen_month = tk.StringVar()
choose_month = ttk.OptionMenu(top_frame, chosen_month, current_month, *months_cz)
choose_month.configure(width=8)

progress = ttk.Progressbar(top_frame, length=147, mode='determinate', orient='vertical')

txt = ScrolledText(top_frame, width=2, height=9)
help_btn = ttk.Button(top_frame, width=1, text='Zobrazit napovedu', command=lambda: show_banner(txt))

show_banner(txt)

start_btn = ttk.Button(top_frame, width=1, text='Start', state='disabled',
                       command=lambda: amn(chosen_month.get(), txt))

btn_opts = {'sticky': 'nswe', 'pady': 5}

choose_month.grid(row=1, column=0, **btn_opts)
progress.grid(row=0, column=1, **btn_opts, rowspan=3, padx=5)
start_btn.grid(row=0, column=0, **btn_opts)
help_btn.grid(row=2, column=0, **btn_opts)
txt.grid(row=0, column=2, rowspan=3, columnspan=2, **btn_opts)

top_frame.grid(row=0, column=0, **opts, pady=10)

bottom_frame = ttk.Frame(root, style="Frames.TFrame")
bottom_frame.grid(row=1, column=0, **opts)

tabs = ttk.Notebook(bottom_frame, width=680)
tabs.grid(row=0, column=0)

# companies = last_data.keys()
for cn in last_data:
    make_frame(cn, 'end')

new_frame = ttk.Frame(tabs)
new_company_name_entry = ttk.Entry(new_frame)
new_company_name_entry.grid(row=0, column=0)
new_company_name_btn = ttk.Button(
    new_frame,
    text='Pridat',
    width=20,
    command=lambda: new_company(len(last_data))
)

new_company_name_btn.grid(row=0, column=1)
tabs.add(new_frame, text='+')

root.mainloop()
